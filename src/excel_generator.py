import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import io

def apply_header_style(cell):
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border

def apply_content_style(cell, is_version_col=False):
    cell.alignment = Alignment(vertical="middle", wrap_text=True, horizontal="center" if is_version_col else "left")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border
    # Padding is not directly supported, rely on wrap_text and column width/row height

def adjust_column_width_and_row_height(ws):
    for col_idx, column_cells in enumerate(ws.columns):
        max_length = 0
        column_letter = get_column_letter(col_idx + 1)
        for cell in column_cells:
            if cell.value:
                cell_length = 0
                if isinstance(cell.value, str):
                    # Consider line breaks for height, max line length for width
                    lines = cell.value.split('\n')
                    max_line_len = max(len(line) for line in lines) if lines else 0
                    cell_length = max_line_len

                elif isinstance(cell.value, (int, float)):
                    cell_length = len(str(cell.value))
                
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = (max_length + 5) # Add some padding
        if adjusted_width > 60: # Max width cap
             adjusted_width = 60
        if adjusted_width < 10: # Min width for version numbers etc.
            if column_letter == 'A' and "Version" in str(ws['A1'].value): # Heuristic for Version #
                 adjusted_width = 10
            elif max_length > 0 : # if there is content
                 adjusted_width = max_length + 5 if max_length + 5 > 10 else 10


        ws.column_dimensions[column_letter].width = adjusted_width

    for row_idx, row_cells in enumerate(ws.rows):
        max_lines = 1
        # Estimate row height based on wrapped text; this is imperfect
        # openpyxl doesn't have a perfect auto-fit row height for wrapped text
        # This is a basic estimation
        for cell in row_cells:
            if cell.value and isinstance(cell.value, str):
                lines = cell.value.count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        if max_lines > 1:
             ws.row_dimensions[row_idx + 1].height = max_lines * 15 # Approx 15 pixels per line


def create_excel_file(ad_data, company_name, links):
    """
    Creates an Excel file in memory with multiple sheets for ad content.
    ad_data is a dictionary where keys are sheet names (e.g., 'email')
    and values are lists of dictionaries (rows).
    links is a dictionary of user-provided links.
    """
    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Email Page
    if 'email' in ad_data and ad_data['email']:
        ws_email = wb.create_sheet("Email")
        email_df = pd.DataFrame(ad_data['email'])
        # Ensure correct column order
        email_cols = ["Version #", "Objective", "Headline", "Subject Line", "Body", "CTA"]
        email_df = email_df.rename(columns={"version": "Version #", "objective_type": "Objective", 
                                            "headline": "Headline", "subject_line": "Subject Line", 
                                            "body": "Body", "cta": "CTA"})
        email_df["Objective"] = "Demand Capture" # As per spec
        
        # Replace placeholder with actual link
        active_lead_link = links.get('active_lead_objective_link', '')
        email_df["Body"] = email_df["Body"].str.replace("[LEAD_OBJECTIVE_LINK]", active_lead_link, regex=False)

        email_df = email_df[email_cols] # Reorder/select columns

        for r_idx, row in enumerate(dataframe_to_rows(email_df, index=False, header=True)):
            ws_email.append(row)
        
        for cell in ws_email[1]: apply_header_style(cell) # Header row
        for row_idx, row in enumerate(ws_email.iter_rows(min_row=2)): # Content rows
            for cell_idx, cell in enumerate(row):
                is_version_col = ws_email.cell(row=1, column=cell_idx+1).value == "Version #"
                apply_content_style(cell, is_version_col)
        adjust_column_width_and_row_height(ws_email)


    # LinkedIn Page
    if 'linkedin' in ad_data and ad_data['linkedin']:
        ws_linkedin = wb.create_sheet("LinkedIn")
        linkedin_df = pd.DataFrame(ad_data['linkedin'])
        linkedin_cols = ["Version #", "Ad Name", "Objective", "Introductory Text", "Image Copy", "Headline", "Destination", "CTA Button"]
        linkedin_df = linkedin_df.rename(columns={
            "version": "Version #", "ad_name": "Ad Name", "objective_type": "Objective",
            "introductory_text": "Introductory Text", "image_copy": "Image Copy",
            "headline": "Headline", "destination_link": "Destination", "cta_button": "CTA Button"
        })
        linkedin_df = linkedin_df[linkedin_cols]

        for r_idx, row in enumerate(dataframe_to_rows(linkedin_df, index=False, header=True)):
            ws_linkedin.append(row)

        for cell in ws_linkedin[1]: apply_header_style(cell)
        for row_idx, row in enumerate(ws_linkedin.iter_rows(min_row=2)):
            for cell_idx, cell in enumerate(row):
                is_version_col = ws_linkedin.cell(row=1, column=cell_idx+1).value == "Version #"
                apply_content_style(cell, is_version_col)
        adjust_column_width_and_row_height(ws_linkedin)

    # Facebook Page
    if 'facebook' in ad_data and ad_data['facebook']:
        ws_facebook = wb.create_sheet("Facebook")
        facebook_df = pd.DataFrame(ad_data['facebook'])
        facebook_cols = ["Version #", "Ad Name", "Objective", "Primary Text", "Image Copy", "Headline", "Link Description", "Destination", "CTA Button"]
        facebook_df = facebook_df.rename(columns={
            "version": "Version #", "ad_name": "Ad Name", "objective_type": "Objective",
            "primary_text": "Primary Text", "image_copy": "Image Copy",
            "headline": "Headline", "link_description": "Link Description",
            "destination_link": "Destination", "cta_button": "CTA Button"
        })
         # Ensure Link Description column exists even if AI doesn't always provide it
        if "Link Description" not in facebook_df.columns:
            facebook_df["Link Description"] = ""
            
        facebook_df = facebook_df[facebook_cols]

        for r_idx, row in enumerate(dataframe_to_rows(facebook_df, index=False, header=True)):
            ws_facebook.append(row)

        for cell in ws_facebook[1]: apply_header_style(cell)
        for row_idx, row in enumerate(ws_facebook.iter_rows(min_row=2)):
            for cell_idx, cell in enumerate(row):
                is_version_col = ws_facebook.cell(row=1, column=cell_idx+1).value == "Version #"
                apply_content_style(cell, is_version_col)
        adjust_column_width_and_row_height(ws_facebook)

    # Google Search Page
    if 'google_search' in ad_data and ad_data['google_search']:
        ws_gsearch = wb.create_sheet("Google Search")
        # Data is expected as {"headlines": [...], "descriptions": [...]}
        gsearch_data = ad_data['google_search']
        max_len = max(len(gsearch_data.get("headlines", [])), len(gsearch_data.get("descriptions", [])))
        
        # Pad shorter list to make DataFrame creation easier
        headlines = gsearch_data.get("headlines", []) + [""] * (max_len - len(gsearch_data.get("headlines", [])))
        descriptions = gsearch_data.get("descriptions", []) + [""] * (max_len - len(gsearch_data.get("descriptions", [])))

        gsearch_df = pd.DataFrame({"Headline": headlines, "Description": descriptions})
        
        for r_idx, row in enumerate(dataframe_to_rows(gsearch_df, index=False, header=True)):
            ws_gsearch.append(row)

        for cell in ws_gsearch[1]: apply_header_style(cell)
        for row in ws_gsearch.iter_rows(min_row=2):
            for cell in row: apply_content_style(cell)
        adjust_column_width_and_row_height(ws_gsearch)

    # Google Display Page
    if 'google_display' in ad_data and ad_data['google_display']:
        ws_gdisplay = wb.create_sheet("Google Display")
        gdisplay_data = ad_data['google_display']
        max_len_display = max(len(gdisplay_data.get("headlines", [])), len(gdisplay_data.get("descriptions", [])))

        headlines_display = gdisplay_data.get("headlines", []) + [""] * (max_len_display - len(gdisplay_data.get("headlines", [])))
        descriptions_display = gdisplay_data.get("descriptions", []) + [""] * (max_len_display - len(gdisplay_data.get("descriptions", [])))

        gdisplay_df = pd.DataFrame({"Headline": headlines_display, "Description": descriptions_display})

        for r_idx, row in enumerate(dataframe_to_rows(gdisplay_df, index=False, header=True)):
            ws_gdisplay.append(row)

        for cell in ws_gdisplay[1]: apply_header_style(cell)
        for row in ws_gdisplay.iter_rows(min_row=2):
            for cell in row: apply_content_style(cell)
        adjust_column_width_and_row_height(ws_gdisplay)

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream